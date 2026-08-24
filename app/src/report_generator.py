from openpyxl import Workbook
from openpyxl.styles import Font

from app.src.database import get_connection


def export_attendance_excel():

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            student_id,
            name,
            attendance_date,
            attendance_time,
            status
        FROM attendance
        ORDER BY attendance_date DESC,
                 attendance_time DESC
    """)

    records = cursor.fetchall()

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Attendance Report"

    headers = [
        "Student ID",
        "Name",
        "Date",
        "Time",
        "Status"
    ]

    for column, header in enumerate(headers, start=1):

        cell = sheet.cell(row=1, column=column)

        cell.value = header

        cell.font = Font(bold=True)

    row_number = 2

    for record in records:

        sheet.cell(row=row_number, column=1).value = record["student_id"]
        sheet.cell(row=row_number, column=2).value = record["name"]
        sheet.cell(row=row_number, column=3).value = str(record["attendance_date"])
        sheet.cell(row=row_number, column=4).value = str(record["attendance_time"])
        sheet.cell(row=row_number, column=5).value = record["status"]

        row_number += 1

    filename = "reports/attendance_report.xlsx"

    workbook.save(filename)

    cursor.close()
    conn.close()

    return filename